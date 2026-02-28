import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

CSV_PATH = "results/summary.csv"
XLSX_PATH = "results/summary_pretty.xlsx"

def autosize_columns(ws, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))

def main():
    df = pd.read_csv(CSV_PATH)
    df.to_excel(XLSX_PATH, index=False, sheet_name="Summary")

    wb = load_workbook(XLSX_PATH)
    ws = wb["Summary"]

    # Freeze header
    ws.freeze_panes = "A2"

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add filter + table style
    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="ResultsTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Number formats for common columns (если есть)
    col_map = {ws.cell(row=1, column=c).value: c for c in range(1, last_col + 1)}

    def set_fmt(col_name, fmt):
        if col_name in col_map:
            c = col_map[col_name]
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=c).number_format = fmt

    set_fmt("U_monthly", "0.00")
    set_fmt("scale", "0.00")
    set_fmt("acc_rag_sim", "0.000")
    set_fmt("acc_ft_sim", "0.000")
    set_fmt("cost_total_rag", "0.0000")
    set_fmt("cost_total_ft", "0.0000")
    set_fmt("cost_per_query_rag", "0.000000")
    set_fmt("cost_per_query_ft", "0.000000")
    set_fmt("rag_index_cost_total", "0.00")
    set_fmt("ft_train_cost_total", "0.00")
    set_fmt("lat_p50_rag_ms", "0.0")
    set_fmt("lat_p95_rag_ms", "0.0")
    set_fmt("lat_p50_ft_ms", "0.0")
    set_fmt("lat_p95_ft_ms", "0.0")
    set_fmt("N_star_break_even", "0")

    # Conditional formatting: accuracy (red->yellow->green)
    for name in ["acc_rag_sim", "acc_ft_sim"]:
        if name in col_map:
            c = col_map[name]
            rng = f"{get_column_letter(c)}2:{get_column_letter(c)}{last_row}"
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(start_type="num", start_value=0.6, start_color="F8696B",
                               mid_type="num", mid_value=0.8, mid_color="FFEB84",
                               end_type="num", end_value=0.95, end_color="63BE7B")
            )

    autosize_columns(ws)
    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")

if __name__ == "__main__":
    main()